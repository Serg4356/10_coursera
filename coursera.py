import requests
import openpyxl
from lxml import etree
from bs4 import BeautifulSoup
import sys
import os
import argparse


def get_courses_link_list(courses_number, http_response):
    etree_courses = etree.fromstring(http_response)
    courses_link_list = []
    for child in etree_courses.getchildren():
        courses_link_list.append(child.getchildren()[0].text)
    return courses_link_list[-courses_number:]


def get_course_info(http_response):
    soup = BeautifulSoup(http_response, 'html.parser')
    course_info = {}
    course_info['name'] = soup.h1.text
    try:
        course_info['average grade'] = soup.select_one(
            'div.ratings-text'
        ).span.text
    except AttributeError:
        course_info['average_grade'] = None
    course_info['weeks required'] = len(
        soup.find_all('div', attrs={'class': 'week'})
    )
    course_info['language'] = soup.select_one('div.rc-Language').text
    course_info['start'] = soup.select_one('div.rc-StartDateString').text
    return course_info


def output_courses_info_to_xls(courses_info, course_attr_names):
    courses_workbook = openpyxl.Workbook()
    work_sheet = courses_workbook.create_sheet('coursera_courses')
    for column, course_attr_name in enumerate(course_attr_names, start=2):
        work_sheet.cell(row=1, column=column).value = course_attr_name
    for row, course in enumerate(courses_info, start=2):
        for column, key in enumerate(course.keys(), start=2):
            work_sheet.cell(row=row, column=1).value = row
            work_sheet.cell(row=row, column=column).value = course[key]
    return courses_workbook


def fetch_http_response(link):
    return requests.get(link).content


def create_parser():
    parser = argparse.ArgumentParser(
        description='Programm searches for courses '
                    'information on coursera.org, and '
                    'outputs them into Excel file'
    )
    parser.add_argument(
        '-d',
        '--display',
        type=bool,
        default=False,
        help='True to display parsing result'
    )
    parser.add_argument(
        '-o',
        '--output',
        default='',
        help='Path to folder'
    )
    return parser


def output_courses_to_console(course_info, output_need):
    print('\n')
    if output_need:
        for course_attr in course_info:
            print('{} - {}'.format(
                course_attr,
                course_info[course_attr])
            )


if __name__ == '__main__':
    parser = create_parser()
    courses_number = 20
    course_attr_names = [
        'name',
        'average grade',
        'weeks required',
        'language',
        'start'
    ]
    coursera_links_xml = fetch_http_response(
        'https://www.coursera.org/sitemap~www~courses.xml'
    )
    courses_link_list = get_courses_link_list(
        courses_number,
        coursera_links_xml
    )
    courses_info = []
    excel_output = parser.parse_args().output
    for course in courses_link_list:
        course_page = fetch_http_response(course)
        course_info = get_course_info(course_page)
        output_courses_to_console(
            course_info,
            excel_output
        )
        courses_info.append(course_info)
    try:
        courses_book = output_courses_info_to_xls(
            courses_info,
            course_attr_names)
        courses_book.save(
            os.path.join(
                parser.parse_args().output,
                'courses.xls')
        )
    except IndexError:
        print('Error! Input path to folder.')
    except PermissionError:
        print('Error! Please, close target excel file')
